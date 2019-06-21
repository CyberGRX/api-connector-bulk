#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#

import os
import json
import requests
from openpyxl import Workbook
from tqdm import tqdm
from utils import sheet_writer
from glom import glom, Coalesce

THIRD_PARTY_TABLE = "Third Parties"
GAPS_TABLE = "Control Gaps (Findings)"
CONTROL_SCORES = "Control Scores"

TP_COLUMNS = [
    ["Company Name", "name"],
    ["Company URL", "url"],

    ["Likelihood", "likelihood_label"],
    ["Likelihood Value", "likelihood_score"],
    ["Impact", "impact_label"],
    ["Impact Value", "impact_score"],

    ["Assessment State", "assessment_status"],
    ["Assessment Progress", "assessment_progress"],
    
    ["Report order status", "subscription_status"],
    ["Report tier", "subscription_tier"],
    ["Report available", "subscription_available"],

    ["Industry", "industry"],
    ["Tags", "tags"],
]    

TP_MAPPING = {
    "name": "name",
    "url": "primary_url",

    "likelihood_label": Coalesce("inherent_risk.likelihood_label", default=None),
    "likelihood_score": Coalesce("inherent_risk.likelihood_score", default=None),
    "impact_label": Coalesce("inherent_risk.impact_label", default=None),
    "impact_score": Coalesce("inherent_risk.impact_score", default=None),

    "assessment_status": Coalesce("assessment.status", default=None),
    "assessment_progress": Coalesce("assessment.progress", default=None),

    "subscription_status": Coalesce("subscription.status", default=None),
    "subscription_tier": Coalesce("subscription.tier", default=None),
    "subscription_available": Coalesce("subscription.is_report_available", default=None),

    "industry": "industry",
    "tags": ("tags", ",".join),
}

GAPS_COLUMNS = [
    ["Company Name", "company_name"],
    ["Control Name", "name"],
    ["Control Number", "number"],
    ["Level", "impact_level"],
    ["Remedy", "remedy"],
]    

GAPS_MAPPING = {
    "company_name": "company_name",
    "name": "name",
    "number": "number",
    "impact_level": "impact_level",
    "remedy": "remedy",
}

SCORE_COLUMNS = [
    ["Company Name", "company_name"],
    ["Control Name", "name"],
    ["Control Number", "number"],
    ["Answer State", "answer_state"],
    ["Effectiveness Score", "effectiveness_score"],
    ["Coverage Score", "coverage_score"],
    ["Maturity Score", "maturity_score"],
]    

SCORE_MAPPING = {
    "company_name": "company_name",
    "name": "name",
    "number": "number",
    "effectiveness_score": Coalesce("effectiveness_score", default=None),
    "coverage_score": Coalesce("coverage_score", default=None),
    "maturity_score": Coalesce("maturity_score", default=None),
    "answer_state": Coalesce("answer_state", default=None),
}

def retireve_ecosystem(): 
    api = os.environ.get('CYBERGRX_BULK_API', "http://127.0.0.1:8080").rstrip("/")
    token = os.environ.get('CYBERGRX_API_TOKEN', None)
    if not token:
        raise Exception("The environment variable CYBERGRX_API_TOKEN must be set")

    uri = api + "/v1/third-parties"
    print("Fetching third parties from " + uri + " this can take some time.") 
    response = requests.get(uri, headers={'Authorization': token.strip()})
    result= json.loads(response.content.decode('utf-8'))

    print("Retrieved " + str(len(result)) + " third parties from your ecosystem, building an excel.")

    wb = Workbook()
    wb['Sheet'].title = THIRD_PARTY_TABLE
    wb.create_sheet(GAPS_TABLE)
    wb.create_sheet(CONTROL_SCORES)

    third_party_writer = sheet_writer(wb, THIRD_PARTY_TABLE, TP_COLUMNS, TP_MAPPING)
    findings_writer = sheet_writer(wb, GAPS_TABLE, GAPS_COLUMNS, GAPS_MAPPING)
    scores_writer = sheet_writer(wb, CONTROL_SCORES, SCORE_COLUMNS, SCORE_MAPPING)

    for tp in tqdm(result, total=len(result), desc="Third Party"):
        third_party_writer(tp)

        for finding in glom(tp, Coalesce("residual_risk.findings", default=[])):
            finding["company_name"] = tp["name"]
            findings_writer(finding)

        for score in glom(tp, Coalesce("residual_risk.scores", default=[])):
            score["company_name"] = tp["name"]
            scores_writer(score)
        
    wb.save("ecosytstem.xlsx")

if __name__ == '__main__':
    retireve_ecosystem()